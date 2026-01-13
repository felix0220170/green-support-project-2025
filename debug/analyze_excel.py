import pandas as pd
import openpyxl

# 读取Excel文件
file_path = '../src/green_finance_2025.xlsx'
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

# 获取工作表的基本信息
print(f"工作表名称: {sheet.title}")
print(f"行数: {sheet.max_row}")
print(f"列数: {sheet.max_column}")

# 检查合并单元格
print("\n合并单元格信息:")
for merged_cell in sheet.merged_cells.ranges:
    print(f"  合并范围: {merged_cell.coord}, 值: {sheet[merged_cell.start_cell.coordinate].value}")

# 读取前20行数据，查看数据结构
print("\n前20行数据预览:")
df = pd.read_excel(file_path, sheet_name=0, header=None)
print(df.head(20))

# 检查空值情况
print("\n各列空值数量:")
print(df.isnull().sum())