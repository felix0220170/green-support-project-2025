import pandas as pd
import re
import openpyxl

# 读取Excel文件
print('正在读取Excel文件...')
wb = openpyxl.load_workbook('green_finance_2025.xlsx')
ws = wb.active

# 查找所有包含0190代码的行
print('\n=== 查找所有包含0190代码的行 ===')
found_rows = []
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row[4] and '0190' in str(row[4]):
        found_rows.append((i, row))
        print(f'行号: {i}, A列: {row[0]}, B列: {row[1]}')
        codes = re.findall(r'\d{4}', str(row[4]))
        print(f'  提取的代码: {codes}')

# 查看找到的行的前后几行
print('\n=== 查看找到的行的前后几行 ===')
for i, row in found_rows:
    print(f'\n=== 行号: {i} 的前后5行 ===')
    
    # 查看前5行
    for j in range(max(1, i-5), i):
        prev_row = list(ws.iter_rows(min_row=j, max_row=j, values_only=True))[0]
        print(f'行号: {j}, A列: {prev_row[0]}, B列: {prev_row[1]}, E列: {repr(prev_row[4]) if prev_row[4] is not None else "None"}')
    
    # 查看当前行
    print(f'行号: {i}, A列: {row[0]}, B列: {row[1]}, E列: {repr(row[4])}')
    codes = re.findall(r'\d{4}', str(row[4]))
    print(f'  提取的代码: {codes}')
    
    # 查看后5行
    for j in range(i+1, min(len(ws)+1, i+6)):
        next_row = list(ws.iter_rows(min_row=j, max_row=j, values_only=True))[0]
        print(f'行号: {j}, A列: {next_row[0]}, B列: {next_row[1]}, E列: {repr(next_row[4]) if next_row[4] is not None else "None"}')
