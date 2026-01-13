import openpyxl
import os
import json

# 定义文件路径
EXCEL_FILE = 'green_finance_2025.xlsx'
OUTPUT_FILE = 'output/parsed_excel_rows.json'

# 创建输出目录
os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

# 辅助函数：检查单元格是否在合并范围内
def get_cell_value(sheet, row, col, merged_cells_info):
    """
    获取单元格的值，如果单元格在合并范围内，则返回合并范围的起始单元格值
    """
    for merged_info in merged_cells_info:
        if (merged_info['min_row'] <= row <= merged_info['max_row'] and 
            merged_info['min_col'] <= col <= merged_info['max_col']):
            return merged_info['value']
    # 如果不在合并范围内，直接获取单元格值
    return sheet.cell(row=row, column=col).value

def parse_excel_with_merged_cells():
    """
    解析Excel文件，处理合并单元格问题
    - 填充共用的standard字段
    - 处理断句问题
    - 保留所有行的完整信息
    """
    # 加载Excel文件
    wb = openpyxl.load_workbook(EXCEL_FILE)
    sheet = wb.active
    
    # 获取工作表的基本信息
    print(f"工作表名称: {sheet.title}")
    print(f"行数: {sheet.max_row}")
    print(f"列数: {sheet.max_column}")
    
    # 收集所有合并单元格的信息
    print("\n收集合并单元格信息...")
    merged_cells_info = []
    
    for merged_cell in sheet.merged_cells.ranges:
        start_cell = merged_cell.start_cell
        value = sheet[start_cell.coordinate].value
        merged_cells_info.append({
            'min_row': merged_cell.min_row,
            'max_row': merged_cell.max_row,
            'min_col': merged_cell.min_col,
            'max_col': merged_cell.max_col,
            'value': value
        })
    
    print(f"共收集到 {len(merged_cells_info)} 个合并单元格")
    
    # 第二步：处理行数据，填充空的standard字段
    print("\n处理行数据...")
    rows = []
    last_standard = None
    last_remark = None
    
    for row_idx in range(1, sheet.max_row + 1):
        # 获取当前行的数据，使用get_cell_value函数处理合并单元格
        row_data = {
            'A': get_cell_value(sheet, row_idx, 1, merged_cells_info),
            'B': get_cell_value(sheet, row_idx, 2, merged_cells_info),
            'C': get_cell_value(sheet, row_idx, 3, merged_cells_info),
            'D': get_cell_value(sheet, row_idx, 4, merged_cells_info),
            'E': get_cell_value(sheet, row_idx, 5, merged_cells_info),
            'F': get_cell_value(sheet, row_idx, 6, merged_cells_info),
            'row_num': row_idx
        }
        
        # 检查是否是章节标题行（所有列值相同）
        is_chapter_title = False
        if all(row_data[col] == row_data['A'] for col in ['B', 'C', 'D', 'E']) and row_data['A'] and isinstance(row_data['A'], str):
            # 如果所有列的值都相同，且不是空值，那么这是一个章节标题行
            is_chapter_title = True
        
        # 如果是章节标题行或列头行（row_num=2），直接跳过
        if is_chapter_title or row_idx == 2:
            continue
        
        # 检查是否是续行（B、C、E、F列为空，D列有值）
        is_continuation_row = False
        if (row_data['B'] is None and row_data['C'] is None and 
            row_data['D'] is not None and isinstance(row_data['D'], str) and 
            row_data['E'] is None and row_data['F'] is None):
            is_continuation_row = True
        
        # 处理续行：直接将D列内容拼接到上一行
        if is_continuation_row and rows:
            if rows[-1]['D'] is not None and isinstance(rows[-1]['D'], str):
                # 直接拼接，不做额外检查
                rows[-1]['D'] += row_data['D']
                # 更新last_standard
                last_standard = rows[-1]['D']
            continue
        
        # 处理D列（standard）：如果为空，则使用上一行的值
        if row_data['D'] is None:
            row_data['D'] = last_standard
        else:
            # 检查是否需要与上一行拼接（断句问题）
            if last_standard and isinstance(last_standard, str) and isinstance(row_data['D'], str):
                # 如果上一行的standard不以句号结尾，且当前行的standard不是完整句子的开头
                if not last_standard.strip().endswith('。') and not row_data['D'].strip().startswith(('1.', '2.', '3.', '4.', '5.', '包括', '对于')):
                    # 拼接断句
                    row_data['D'] = last_standard + row_data['D']
            # 更新last_standard
            last_standard = row_data['D']
        
        # 处理E列（remark）：如果不为空，检查是否需要与上一行拼接（断句问题）
        if row_data['E'] is not None:
            # 检查是否需要与上一行拼接
            if rows and rows[-1]['E'] is not None and isinstance(rows[-1]['E'], str) and isinstance(row_data['E'], str):
                # 如果上一行的remark不以句号结尾，且当前行的remark不是完整句子的开头
                if not rows[-1]['E'].strip().endswith('。') and not row_data['E'].strip().startswith(('1.', '2.', '3.', '4.', '5.', '包括', '对于')):
                    # 更新上一行的remark，将当前行的remark合并到上一行
                    rows[-1]['E'] += row_data['E']
                    # 当前行的remark设为空
                    row_data['E'] = None
                else:
                    # 更新last_remark为当前行的remark
                    last_remark = row_data['E']
            else:
                # 更新last_remark为当前行的remark
                last_remark = row_data['E']
        else:
            # 如果当前行的remark为空，保持last_remark不变
            pass
        
        rows.append(row_data)
    
    # 第三步：保存处理后的数据
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)
    
    print(f"\n处理完成！")
    print(f"处理后的数据已保存到: {OUTPUT_FILE}")
    
    # 统计信息
    valid_standard_count = sum(1 for row in rows if row['D'] is not None)
    empty_standard_count = sum(1 for row in rows if row['D'] is None)
    valid_remark_count = sum(1 for row in rows if row['E'] is not None)
    empty_remark_count = sum(1 for row in rows if row['E'] is None)
    
    print(f"\n数据统计:")
    print(f"  总行数: {len(rows)}")
    print(f"  Standard字段有效: {valid_standard_count}")
    print(f"  Standard字段为空: {empty_standard_count}")
    print(f"  Remark字段有效: {valid_remark_count}")
    print(f"  Remark字段为空: {empty_remark_count}")
    
    return rows

if __name__ == "__main__":
    parsed_rows = parse_excel_with_merged_cells()