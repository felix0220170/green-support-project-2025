import pandas as pd
import re
import openpyxl

# 读取Excel文件
wb = openpyxl.load_workbook('green_finance_2025.xlsx')
ws = wb.active

# 查看5.1.11行（行号590-600）的E列内容
print('=== 查看5.1.11行附近的E列内容 (行号590-610) ===')
for i in range(589, 610):  # Excel行号从1开始，Python从0开始，所以实际是590-611行
    row = list(ws.iter_rows(min_row=i+1, max_row=i+1, values_only=True))[0]
    print(f'行号: {i+1}, A列: {row[0]}, E列: {repr(row[4]) if row[4] is not None else "None"}')
    
    # 提取备注中的4位数字代码
    if row[4]:
        codes = re.findall(r'\d{4}', str(row[4]))
        if codes:
            print(f'  提取的代码: {codes}')
            if '0190' in codes or '0399' in codes:
                print(f'  ✓ 找到目标代码: {[c for c in codes if c in ["0190", "0399"]]}')
    print()

# 查看5.1.14行（行号614-630）的E列内容
print('=== 查看5.1.14行附近的E列内容 (行号614-630) ===')
for i in range(613, 630):  # 实际是614-631行
    row = list(ws.iter_rows(min_row=i+1, max_row=i+1, values_only=True))[0]
    print(f'行号: {i+1}, A列: {row[0]}, E列: {repr(row[4]) if row[4] is not None else "None"}')
    
    # 提取备注中的4位数字代码
    if row[4]:
        codes = re.findall(r'\d{4}', str(row[4]))
        if codes:
            print(f'  提取的代码: {codes}')
            if '0190' in codes or '0399' in codes:
                print(f'  ✓ 找到目标代码: {[c for c in codes if c in ["0190", "0399"]]}')
    print()
